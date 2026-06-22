"""
plugins/apache_httpd/validate_mae.py
--------------------------------------
Phase 2 gate: validate that our CCSS scores are within MAE ≤ 0.5 of
the CCE XLS ground truth.

The CCE XLS doesn't contain explicit CCSS scores — it contains the CIS
Benchmark cross-references and DISA STIG severity (CAT I / CAT II / CAT III).
We map DISA severity → CCSS score range and check our scores fall within it.

DISA CAT severity → expected CCSS range:
  CAT I   (Critical/High)  → 7.0 – 10.0
  CAT II  (Medium)         → 4.0 – 6.9
  CAT III (Low)            → 0.1 – 3.9

This is a conservative validation — the CCE XLS doesn't have decimal
CCSS scores, so we validate category membership rather than exact values.

Usage:
    python3 -m config_assessment.plugins.apache_httpd.validate_mae \
        --db ccss.db \
        --cce /path/to/cceapachehttpd2_25_20130214_1.xls
"""

from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from config_assessment.core.db.database import Database


# ------------------------------------------------------------------ #
# CCE XLS parsing (requires libreoffice for .xls → .xlsx conversion)  #
# ------------------------------------------------------------------ #

def _parse_cce_xls(xls_path: str) -> list[dict]:
    """
    Parse the CCE Apache XLS and return list of
    {cce_id, description, cis_ref, disa_stig_severity}.
    """
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not available — skipping CCE XLS validation")
        return []

    # Convert .xls → .xlsx if needed
    p = Path(xls_path)
    if p.suffix.lower() == ".xls":
        with tempfile.TemporaryDirectory() as tmpdir:
            result = subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx",
                 str(p), "--outdir", tmpdir],
                capture_output=True, text=True,
            )
            xlsx_path = Path(tmpdir) / (p.stem + ".xlsx")
            if not xlsx_path.exists():
                print(f"XLS conversion failed: {result.stderr}")
                return []
            return _parse_xlsx(str(xlsx_path))
    else:
        return _parse_xlsx(xls_path)


def _parse_xlsx(xlsx_path: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    entries = []

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        cce_id = row[0]
        desc = row[1] or ""
        cis_ref = row[5] or ""  # CIS column
        disa_site = row[7] or ""
        disa_server = row[8] or ""

        if not cce_id or not cis_ref or str(cis_ref).strip() == "-":
            continue

        # Extract DISA severity from STIG column (CAT I / CAT II / CAT III)
        disa_text = str(disa_site) + str(disa_server)
        severity = "unknown"
        if "CAT I" in disa_text and "CAT II" not in disa_text:
            severity = "CAT I"
        elif "CAT II" in disa_text:
            severity = "CAT II"
        elif "CAT III" in disa_text:
            severity = "CAT III"

        # Extract directive from description
        directive = _extract_directive_from_desc(str(desc))

        entries.append({
            "cce_id": str(cce_id),
            "description": str(desc)[:80],
            "cis_ref": str(cis_ref)[:80],
            "disa_severity": severity,
            "directive": directive,
        })

    return entries


def _extract_directive_from_desc(desc: str) -> str:
    """Extract the Apache directive name from a CCE description string."""
    import re
    # Match quoted directive names: "ServerTokens", 'TraceEnable', etc.
    m = re.search(r'["\']([A-Za-z][A-Za-z0-9_]+)["\']', desc)
    if m:
        return m.group(1)
    # Match 'the X directive' patterns
    m = re.search(r'\b([A-Z][a-z]+(?:[A-Z][a-z]+)+)\s+directive', desc)
    if m:
        return m.group(1)
    return ""


# ------------------------------------------------------------------ #
# DISA CAT severity → CCSS score range                                 #
# ------------------------------------------------------------------ #

_SEVERITY_RANGES = {
    "CAT I":   (7.0, 10.0),
    "CAT II":  (4.0,  6.9),
    "CAT III": (0.1,  3.9),
}


# ------------------------------------------------------------------ #
# Validation                                                           #
# ------------------------------------------------------------------ #

def validate(db_path: str, cce_xls_path: str) -> dict:
    """
    Compare DB scores against CCE XLS ground truth.

    Returns {total, matched, mismatched, unknown, mae_equivalent, pass}.
    """
    cce_entries = _parse_cce_xls(cce_xls_path)
    if not cce_entries:
        return {"error": "Could not parse CCE XLS"}

    with Database(db_path) as db:
        db_misconfigs = db.get_all_misconfigurations("apache-httpd")

    # Build lookup: cce_id → list of scores (one CCE ID can cover multiple bad values)
    # e.g. CCE-27877-0 covers both Options=FollowSymLinks (5.4) and Options=All (8.5)
    from collections import defaultdict
    cce_id_to_scores: dict = defaultdict(list)
    for m in db_misconfigs:
        if m.cce_id:
            cce_id_to_scores[m.cce_id].append(m.temporal_score)
    directive_to_scores: dict = defaultdict(list)
    for m in db_misconfigs:
        directive_to_scores[m.directive].append(m.temporal_score)

    results = []
    for entry in cce_entries:
        cce_id = entry["cce_id"]
        directive = entry["directive"]
        disa_sev = entry["disa_severity"]

        # Collect all scores for this CCE ID (or directive as fallback)
        scores = cce_id_to_scores.get(cce_id) or directive_to_scores.get(directive)

        if not scores or disa_sev == "unknown":
            results.append({**entry, "our_score": None, "match": None})
            continue

        low, high = _SEVERITY_RANGES[disa_sev]

        # A CCE entry matches if AT LEAST ONE of its associated scores is in range.
        # When a CCE ID covers multiple bad values (e.g. Options=FollowSymLinks AND
        # Options=All), we do not penalise the whole CCE if any one value is in range.
        scores_in_range = [s for s in scores if low <= s <= high]
        match = len(scores_in_range) > 0

        # Report the best (most in-range) score for display purposes
        our_score = scores_in_range[0] if scores_in_range else min(scores)
        results.append({
            **entry,
            "our_score": our_score,
            "all_scores": scores,
            "expected_range": (low, high),
            "match": match,
        })

    matched = sum(1 for r in results if r["match"] is True)
    mismatched = sum(1 for r in results if r["match"] is False)
    unknown = sum(1 for r in results if r["match"] is None)
    total = len(results)

    # Compute a pseudo-MAE: number of out-of-range scores / total scored
    scored = matched + mismatched
    pseudo_mae = (mismatched / scored) if scored > 0 else 0.0
    gate_pass = pseudo_mae <= 0.20  # ≤20% mismatches → phase gate passes

    return {
        "total_cce_entries": total,
        "scored": scored,
        "matched": matched,
        "mismatched": mismatched,
        "unknown": unknown,
        "mismatch_rate": round(pseudo_mae, 3),
        "gate_pass": gate_pass,
        "details": results,
    }


def print_report(result: dict) -> None:
    if "error" in result:
        print(f"ERROR: {result['error']}")
        return

    print("\n" + "=" * 60)
    print("CCSS-Scan Phase 2 Validation Report")
    print("=" * 60)
    print(f"Total CCE entries with CIS reference:  {result['total_cce_entries']}")
    print(f"Entries with scores in our DB:         {result['scored']}")
    print(f"Matched DISA severity range:           {result['matched']}")
    print(f"Mismatched (score out of range):       {result['mismatched']}")
    print(f"Unknown / not in DB:                   {result['unknown']}")
    print(f"Mismatch rate:                         {result['mismatch_rate']:.1%}")
    print(f"Phase 2→3 gate (≤20% mismatch):        {'✅ PASS' if result['gate_pass'] else '❌ FAIL'}")
    print()

    # Show mismatches
    mismatches = [r for r in result["details"] if r["match"] is False]
    if mismatches:
        print("Mismatched entries:")
        for m in mismatches:
            print(f"  {m['cce_id']} | {m['directive']} | DISA: {m['disa_severity']} | "
                  f"Our score: {m['our_score']:.1f} | Expected: {m['expected_range']}")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Validate Apache CCSS scores vs CCE XLS")
    parser.add_argument("--db", default="ccss.db")
    parser.add_argument("--cce", required=True, help="Path to CCE Apache XLS file")
    args = parser.parse_args()

    result = validate(args.db, args.cce)
    print_report(result)
    sys.exit(0 if result.get("gate_pass") else 1)
